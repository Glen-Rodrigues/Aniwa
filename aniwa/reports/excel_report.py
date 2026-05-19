from pathlib import Path

from aniwa.models.profile import DatasetProfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def render_excel_report(profile: DatasetProfile, output: str | None = None) -> str:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    _populate_summary_sheet(summary_ws, profile)

    if profile.columns:
        columns_ws = wb.create_sheet("Columns")
        _populate_columns_sheet(columns_ws, profile)

    if profile.quality:
        quality_ws = wb.create_sheet("Data Quality")
        _populate_quality_sheet(quality_ws, profile)

    if profile.insights is not None:
        insights_ws = wb.create_sheet("Insights")
        _populate_insights_sheet(insights_ws, profile)

    if output:
        Path(output).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return output

    return "Excel report generated"


def _format_header_row(ws, column_count: int) -> None:
    header_font = Font(bold=True)
    for col_num in range(1, column_count + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _adjust_column_widths(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def _populate_summary_sheet(ws, profile: DatasetProfile) -> None:
    ws.append(["Metric", "Value"])
    _format_header_row(ws, 2)

    if profile.summary:
        ws.append(["Rows", profile.summary.rows])
        ws.append(["Columns", profile.summary.columns])
    else:
        ws.append(["Summary", "Not available"])

    _adjust_column_widths(ws)


def _populate_columns_sheet(ws, profile: DatasetProfile) -> None:
    headers = [
        "Column",
        "Type",
        "Null Count",
        "Null %",
        "Unique Count",
        "Min",
        "Max",
        "Mean",
        "Median",
        "Std",
    ]
    ws.append(headers)
    _format_header_row(ws, len(headers))

    for column in profile.columns:
        if column.numeric_stats:
            stats = column.numeric_stats
            ws.append(
                [
                    column.name,
                    column.dtype,
                    column.null_count,
                    column.null_percent,
                    column.unique_count,
                    stats.min,
                    stats.max,
                    stats.mean,
                    stats.median,
                    stats.std,
                ]
            )
        else:
            ws.append(
                [
                    column.name,
                    column.dtype,
                    column.null_count,
                    column.null_percent,
                    column.unique_count,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )

    _adjust_column_widths(ws)


def _populate_quality_sheet(ws, profile: DatasetProfile) -> None:
    ws.append(["Metric", "Value"])
    _format_header_row(ws, 2)
    ws.append(["Duplicate Rows", profile.quality.duplicate_rows])
    ws.append(["Duplicate %", profile.quality.duplicate_percent])
    _adjust_column_widths(ws)


def _populate_insights_sheet(ws, profile: DatasetProfile) -> None:
    ws.append(["Level", "Message"])
    _format_header_row(ws, 2)

    if not profile.insights:
        ws.append(["info", "No major issues detected."])
    else:
        for insight in profile.insights:
            ws.append([insight.level, insight.message])

    _adjust_column_widths(ws)
